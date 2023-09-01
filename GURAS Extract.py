#v1 - Script to extract GURAS Address based on ptlotsecpn (Plan Type / Lot no / Section No / Plan No) input from CSV. Results are stored in GURAS_DATA table and exported to CSV file 'GURAS_Results.csv'

print("Importing Python Packages..")
import time
import cx_Oracle
import config
#import SQLGURAS
import requests
#import urllib
import json
import string
import pandas as pd
import openpyxl
from datetime import date, datetime
import sys
print("All packages imported Successfully!\n")

today = date.today()

#Handle Null Values
def ifnull(var, val):
	if var is None:
		return val
	elif pd.isna(var):
		return val
	return var

#Get Next ID
def getNextId(column: str, table: str) -> int:
	c.execute("select max({}) from {}".format(column, table))
	result = c.fetchone()
	
	#If records exist, increment next id, else start at 1
	if result[0] != None:
		nextId = result[0] + 1
	else:
		nextId = 1
	
	return nextId

def connectDB():
	#Connects to GPR Database
	connection = None

	oc_attempts = 0

	while oc_attempts < 2:
		if oc_attempts == 0:
			print("Trying DPE IP: {}".format(config.dsnDPE))
			dsn = config.dsnDPE
		else:
			dsn = config.dsnDCS
			print("Trying DCS IP: {}".format(config.dsnDCS))
			
		try:
			connection = cx_Oracle.connect(
				config.username,
				config.password,
				dsn,
				encoding=config.encoding)

			# show the version of the Oracle Database
			print(connection.version," Connection Successful!")
			oc_attempts = 2
		except cx_Oracle.Error as error:
			print(error)
			oc_attempts += 1
			
	return connection

if __name__ == "__main__":
	
	print("User name: {}".format(sys.argv[1]))
	
	#Set filename for results exported
	exportFile = "gurasResults_{}.xlsx".format(sys.argv[1])
	
	#Prompt User to run in Table mode or CSV only mode
	while True:
		mode = int(input("\nPlease select mode for GURAS extraction:\n\t[1] - Save to table and csv\n\t[2] - Save to csv only\nEnter 1 or 2: "))
		if mode not in [1,2]:
			print("Invalid input, please enter 1 or 2!")
			continue
		else:
			break

	if mode == 1:
		print("\nTable and csv Mode selected!\n")
	elif mode == 2:
		print("\nCsv only Mode selected!\n")
			
	#If in table mode, check connection is successful before continuing
	if mode == 1:
		connection = connectDB()
		
		#release the connection
		if connection:
			connection.close()
		else:
			sys.exit("\nERROR: Connection to GPR Database unsuccessful, unable to continue process...")

	#Read 'Lots_to_Guras.csv' file to dataframe
	print("------------------------------------------------")
	print("Reading csv file 'lots_to_guras.csv'...")
	print("------------------------------------------------")
	
	in_csv_path = "lots_to_guras.csv"
	
	df_input_unduped = pd.read_csv(in_csv_path)
	
	df_input_duped = df_input_unduped.drop_duplicates()
	
	df_input = df_input_duped.reset_index()
	
	#Get column headers
	column_headers = list(df_input.columns.values)
	
	#Split into individual columns
	df_split = pd.DataFrame(columns=['ptlotsecpn','lot_no','section_no','plan_type','plan_no'])
	
	df_split[['plan_type','lot_no','section_no','plan_no']] = df_input[column_headers[1]].str.split('/', expand=True)
	
	#df_split['ptlotsecpn'] = df_input['ptlotsecpn']
	df_split['ptlotsecpn'] = df_input[column_headers[1]]
	
	#Initialise string to store ptlotsecpn for PropID query
	lotstring = ''
	
	#Initialise List to store all Json results
	propidResults = list()
	
	#Set query URL
	queryURL = config.GURASPropLot
	
	len_lot = len(df_split)
	
	#Go through lots and extract PROPID
	for i, row in df_split.iterrows():
		if lotstring == '':
			lotstring += "'{}'".format(row["ptlotsecpn"])
		else:
			lotstring += ",'{}'".format(row["ptlotsecpn"])

		if (i + 1) % 200 == 0 or (i + 1) == len(df_split):

			params = {
				'f':'json',
				'returnGeometry':'false',
				'outSR':'4326',
				'OutFields':'ptlotsecpn,propid,sppropid',
				'where':'ptlotsecpn in ({})'.format(lotstring)
			}
			retries = 0
			success = False
			while not success:
				try:
					response = requests.get(url=queryURL, params=params)
					success = True
				except requests.exceptions.RequestException as e:
					print(e)
					retries += 1
					if retries > 9:
						while True:
							select = input("\nRequest to Lot-PropID service failed 10 times, Do you want to try again? y/n\n")
							if select == "y":
								retries = 0
								break
							elif select == "n":
								print("GURAS Extract process Aborted!!")
								sys.exit()
							else:
								print("Invalid selection. Please enter y or n")
				
				while response.status_code != 200 and success:
					print("Response code: {}".format(response.status_code))
					select2 = input("\nInvalid response received, run query again? y/n\n")
					if select2 == "y":
						retries = 0
						success = False
						break
					elif select2 == "n":
						print("GURAS Extract process Aborted!!")
						sys.exit()
					else:
						print("Invalid selection. Please enter y or n")
						
			jsonResult = json.loads(response.text)
					
			if jsonResult.get('features'):
				#iterate through all features in JSON response and add to Result list
				for jr in range(len(jsonResult['features'])):
					propidResults.append(jsonResult['features'][jr])
			
			#Reset
			lotstring = ''
			
		#Display Progress
		print("{}/{} lots processed...".format(i + 1,len_lot), end="\r")
	print("{} lots processed        ".format(i + 1))
	
	print("Transforming to final results...")
	print("------------------------------------------------")

	#Store Final PropID results to dataframe to transform into final results
	df_propidResults = pd.json_normalize(propidResults)
	
	df_propidFinal = pd.merge(left=df_split, right=df_propidResults, left_on='ptlotsecpn', right_on='attributes.ptlotsecpn')
	
	#Export Prop ID Results to Excel file
	print("Saving PropID Results to Excel file: {}".format(exportFile))
	print("------------------------------------------------")
	
	wb = openpyxl.Workbook(write_only=True)
	sh = wb.create_sheet("lot_propid")
	
	#Set headers
	sh.append(['PROPID','SPPROPID','UNIQUE_PROPID','PTLOTSECPN','LOT_NO','SECTION_NO','PLAN_TYPE','PLAN_NO'])
	
	for i, row in df_propidFinal.iterrows():
		if pd.notna(row['attributes.sppropid']):
			unique_propid = row['attributes.sppropid']
		else:
			unique_propid = row['attributes.propid']
			
		sh.append([row['attributes.propid'],row['attributes.sppropid'],unique_propid,row['ptlotsecpn'],row['lot_no'],row['section_no'],row['plan_type'],row['plan_no']])
	wb.save(exportFile)
	
	#GURAS EXTRACT
	print("Extracting GURAS Address data...")
	print("------------------------------------------------")
	
	#Remove propid duplicates
	df_propid_dedupe = df_propidFinal.drop_duplicates(subset=['attributes.propid'])
	
	df_propid_to_guras = df_propid_dedupe.reset_index()
	
	#Initialise string to store PropID for GURAS query
	pstring = ''
	
	#Initialise List to store all Json results
	gurasResults = list()
	
	#Set query URL
	queryURL = config.GURASAddress
	queryWhere = ''
	
	len_prop = len(df_propid_to_guras)
	
	for i, row in df_propid_to_guras.iterrows():
		if (pd.notna(row["attributes.propid"])):
			if pstring == '':
				pstring += "'{}'".format(int(row["attributes.propid"]))
			else:
				pstring += ",'{}'".format(int(row["attributes.propid"]))
		
		#Every 200 propIDs query services
		if (i + 1) % 200 == 0 or (i + 1) == len(df_propid_to_guras):
			queryWhere = "(propid in (" + pstring + ") and principaladdresstype = 1)"
			params = {
				'f':'json',
				'returnGeometry':'false',
				'outSR':'4326',
				'OutFields':'*',
				'where':queryWhere
			}
			retries = 0
			success = False
			while not success:
				try:
					response = requests.get(url=queryURL, params=params)
					success = True
				except requests.exceptions.RequestException as e:
					print(e)
					retries += 1
					if retries > 9:
						while True:
							select = input("\nRequest to GURAS service failed 10 times, Do you want to try again? y/n\n")
							if select == "y":
								retries = 0
								break
							elif select == "n":
								print("GURAS Extract process Aborted!!")
								sys.exit()
							else:
								print("Invalid selection. Please enter y or n")
				
				while response.status_code != 200 and success:
					print("Response code: {}".format(response.status_code))
					select2 = input("\nInvalid response received, run query again? y/n\n")
					if select2 == "y":
						retries = 0
						success = False
						break
					elif select2 == "n":
						print("GURAS Extract process Aborted!!")
						sys.exit()
					else:
						print("Invalid selection. Please enter y or n")
						
			jsonResult = json.loads(response.text)
			
			if jsonResult.get('features'):
				#iterate through all features in JSON response and add to Result list
				for jr in range(len(jsonResult['features'])):
					gurasResults.append(jsonResult['features'][jr])
			#Reset
			pstring = ''
			
		#Display Progress
		print("{}/{} propid's processed...".format(i + 1,len_prop), end="\r")
	
	print("{} propid's processed      ".format(i + 1))
		
	print("Transforming Address results...")
	print("------------------------------------------------")
	
	#Store Final GURAS results to dataframe to transform into final results
	df_gurasResults = pd.json_normalize(gurasResults)
	
	#Merge address data
	m_description = list()
	m_address = list()
	m_suburb = list()
	
	for i, grow in df_gurasResults.iterrows():
		description = ''
		address = ''
		suburb = ''
		
		#Location Description
		if pd.notna(grow['attributes.locationdescription']):
			description = "{}".format(grow['attributes.locationdescription'])

		#Address Site Name
		if pd.notna(grow['attributes.addresssitename']):
			if len(description) > 0:
				description = "{}, ".format(description)
			description = "{}{}".format(description,grow['attributes.addresssitename'])
	
		#Building Name
		if pd.notna(grow['attributes.buildingname']):
			if len(description) > 0:
				description = "{}, ".format(description)
			description = "{}{}".format(description,grow['attributes.buildingname'])

		#Level
		if pd.notna(grow['attributes.leveltype']):
			address = "{} ".format(grow['attributes.leveltype'])
		
		#Check for Nonetype values
		if pd.notna(grow['attributes.levelnumberprefix']):
			levelnumberprefix = grow['attributes.levelnumberprefix']
		else:
			levelnumberprefix = ''
			
		if pd.notna(grow['attributes.levelnumber']):
			levelnumber = grow['attributes.levelnumber']
		else:
			levelnumber = ''
			
		if pd.notna(grow['attributes.levelnumbersuffix']):
			levelnumbersuffix = grow['attributes.levelnumbersuffix']
		else:
			levelnumbersuffix = ''
		
		address = "{}{}{}{}".format(address,levelnumberprefix,levelnumber,levelnumbersuffix)
		
		if pd.notna(grow['attributes.levelnumberprefix']) or pd.notna(grow['attributes.levelnumber']) or pd.notna(grow['attributes.levelnumbersuffix']):
			address = "{}, ".format(address)
			
		#Unit
		if pd.notna(grow['attributes.unittype']):
			address = "{}{} ".format(address, grow['attributes.unittype'])
		
		#Check for Nonetype values
		if pd.notna(grow['attributes.unitnumberprefix']):
			unitnumberprefix = grow['attributes.unitnumberprefix']
		else:
			unitnumberprefix = ''
			
		if pd.notna(grow['attributes.unitnumber']):
			unitnumber = grow['attributes.unitnumber']
		else:
			unitnumber = ''
			
		if pd.notna(grow['attributes.unitnumbersuffix']):
			unitnumbersuffix = grow['attributes.unitnumbersuffix']
		else:
			unitnumbersuffix = ''
		
		address = "{}{}{}{}".format(address,unitnumberprefix,unitnumber,unitnumbersuffix)
		
		if pd.notna(grow['attributes.unitnumberprefix']) or pd.notna(grow['attributes.unitnumber']) or pd.notna(grow['attributes.unitnumbersuffix']):
			address = "{}/".format(address)
		
		#NUMBER AND STREET NAMES
		
		#First Road
		#Check for Nonetype values
		if pd.notna(grow['attributes.housenumberfirstprefix']):
			housenumberfirstprefix = grow['attributes.housenumberfirstprefix']
		else:
			housenumberfirstprefix = ''
			
		if pd.notna(grow['attributes.housenumberfirst']):
			housenumberfirst = int(grow['attributes.housenumberfirst'])
		else:
			housenumberfirst = ''
			
		if pd.notna(grow['attributes.housenumberfirstsuffix']):
			housenumberfirstsuffix = grow['attributes.housenumberfirstsuffix']
		else:
			housenumberfirstsuffix = ''
			
		address = "{}{}{}{}".format(address,housenumberfirstprefix,housenumberfirst,housenumberfirstsuffix)
		
		#Check for Nonetype values
		if pd.notna(grow['attributes.housenumbersecondprefix']):
			housenumbersecondprefix = grow['attributes.housenumbersecondprefix']
		else:
			housenumbersecondprefix = ''
			
		if pd.notna(grow['attributes.housenumbersecond']):
			housenumbersecond = int(grow['attributes.housenumbersecond'])
		else:
			housenumbersecond = ''
			
		if pd.notna(grow['attributes.housenumbersecondsuffix']):
			housenumbersecondsuffix = grow['attributes.housenumbersecondsuffix']
		else:
			housenumbersecondsuffix = ''
		
		#If at least one value from first and second house number fields have value, include '-'
		if (pd.notna(grow['attributes.housenumberfirstprefix']) or pd.notna(grow['attributes.housenumberfirst']) or pd.notna(grow['attributes.housenumberfirstsuffix'])) and (pd.notna(grow['attributes.housenumbersecondprefix']) or pd.notna(grow['attributes.housenumbersecond']) or pd.notna(grow['attributes.housenumbersecondsuffix'])):
			address = "{}-".format(address)
		
		address = "{}{}{}{}".format(address,housenumbersecondprefix,housenumbersecond,housenumbersecondsuffix)
		
		if pd.notna(grow['attributes.housenumberfirstprefix']) or pd.notna(grow['attributes.housenumberfirst']) or pd.notna(grow['attributes.housenumberfirstsuffix']) or pd.notna(grow['attributes.housenumbersecondprefix']) or pd.notna(grow['attributes.housenumbersecond']) or pd.notna(grow['attributes.housenumbersecondsuffix']):
			address = "{} ".format(address)
		
		#Check for Nonetype values
		if pd.notna(grow['attributes.roadname']):
			roadname = string.capwords(grow['attributes.roadname'])
		else:
			roadname = ''
		
		if pd.notna(grow['attributes.roadname']):
			address = "{}{} ".format(address,roadname)
			
		if pd.notna(grow['attributes.roadtype']):
			roadtype = string.capwords(grow['attributes.roadtype'])
		else:
			roadtype = ''
			
		address = "{}{}".format(address,roadtype)
			
		if pd.notna(grow['attributes.roadsuffix']):
			roadsuffix = string.capwords(grow['attributes.roadsuffix'])
		else:
			roadsuffix = ''
			
		if pd.notna(grow['attributes.roadsuffix']):
			address = "{} {}".format(address, roadsuffix)
			
		#Second Road
		if pd.notna(grow['attributes.secondroadname']) or pd.notna(grow['attributes.secondroadsuffix']) or pd.notna(grow['attributes.secondroadtype']):
			address = "{}/".format(address)
		
		#Check for Nonetype values
		if pd.notna(grow['attributes.secondroadname']):
			secondroadname = string.capwords(grow['attributes.secondroadname'])
		else:
			secondroadname = ''
			
		if pd.notna(grow['attributes.secondroadsuffix']):
			secondroadsuffix = string.capwords(grow['attributes.secondroadsuffix'])
		else:
			secondroadsuffix = ''
			
		if pd.notna(grow['attributes.secondroadtype']):
			secondroadtype = string.capwords(grow['attributes.secondroadtype'])
		else:
			secondroadtype = ''
			
		address = "{}{}".format(address,secondroadname)
		
		if pd.notna(grow['attributes.secondroadname']):
			address = "{} ".format(address)
			
		address = "{}{}".format(address,secondroadtype)
		
		if pd.notna(grow['attributes.secondroadtype']):
			address = "{} ".format(address)
			
		address = "{}{}".format(address,secondroadsuffix)
		
		#SUBURB AND POSTCODE
		if pd.notna(grow['attributes.suburbname']):
			suburbname = string.capwords(grow['attributes.suburbname'])
		else:
			suburbname = ''
			
		if pd.notna(grow['attributes.postcode']):
			postcode = grow['attributes.postcode']
		else:
			postcode = ''
		
		suburb = "{} {}".format(suburbname, postcode)
		
		m_description.append(string.capwords(description))
		m_address.append(address)
		m_suburb.append(suburb)
	
	#Add merged Address data columns
	df_gurasResults["merged_property_description"] = m_description
	df_gurasResults["merged_address"] = m_address
	df_gurasResults["merged_suburb"] = m_suburb
	
	#Export GURAS Results to Excel file
	print("Saving GURAS Results to Excel file: {}".format(exportFile))
	print("------------------------------------------------")
	
	wb2 = openpyxl.load_workbook(exportFile)
	sh2 = wb2.create_sheet("guras_address")
	
	#Set headers
	sh2.append(['PROPID','SPPROPID','UNIQUE_PROPID','merged_property_description','merged_address','merged_suburb','objectid','createdate','gurasid','addresstype','ruraladdress','principaladdresstype','addressstringtype','principaladdresssiteoid','officialaddressstringoid','roadside','housenumberfirstprefix','housenumberfirst','housenumberfirstsuffix','housenumbersecondprefix','housenumbersecond','housenumbersecondsuffix','roadname','roadtype','roadsuffix','unittype','unitnumberprefix','unitnumber','unitnumbersuffix','leveltype','levelnumberprefix','levelnumber','levelnumbersuffix','addresssitename','buildingname','locationdescription','privatestreetname','privatestreettype','privatestreetsuffix','secondroadname','secondroadtype','secondroadsuffix','suburbname','state','postcode','council','deliverypointid','deliverypointbarcode','addressconfidence','contributororigin','contributorid','contributoralignment','routeoid','gnafprimarysiteid','containment'])
	
	for i, row in df_gurasResults.iterrows():
		if pd.notna(row['attributes.sppropid']):
			unique_propid = row['attributes.sppropid']
		else:
			unique_propid = row['attributes.propid']
			
		sh2.append([row['attributes.propid'],row['attributes.sppropid'],unique_propid,row['merged_property_description'],row['merged_address'],row['merged_suburb'],row['attributes.objectid'],row['attributes.createdate'],row['attributes.gurasid'],row['attributes.addresstype'],row['attributes.ruraladdress'],row['attributes.principaladdresstype'],row['attributes.addressstringtype'],row['attributes.principaladdresssiteoid'],row['attributes.officialaddressstringoid'],row['attributes.roadside'],row['attributes.housenumberfirstprefix'],row['attributes.housenumberfirst'],row['attributes.housenumberfirstsuffix'],row['attributes.housenumbersecondprefix'],row['attributes.housenumbersecond'],row['attributes.housenumbersecondsuffix'],row['attributes.roadname'],row['attributes.roadtype'],row['attributes.roadsuffix'],row['attributes.unittype'],row['attributes.unitnumberprefix'],row['attributes.unitnumber'],row['attributes.unitnumbersuffix'],row['attributes.leveltype'],row['attributes.levelnumberprefix'],row['attributes.levelnumber'],row['attributes.levelnumbersuffix'],row['attributes.addresssitename'],row['attributes.buildingname'],row['attributes.locationdescription'],row['attributes.privatestreetname'],row['attributes.privatestreettype'],row['attributes.privatestreetsuffix'],row['attributes.secondroadname'],row['attributes.secondroadtype'],row['attributes.secondroadsuffix'],row['attributes.suburbname'],row['attributes.state'],row['attributes.postcode'],row['attributes.council'],row['attributes.deliverypointid'],row['attributes.deliverypointbarcode'],row['attributes.addressconfidence'],row['attributes.contributororigin'],row['attributes.contributorid'],row['attributes.contributoralignment'],row['attributes.routeoid'],row['attributes.gnafprimarysiteid'],row['attributes.containment']])
		
	wb2.save(exportFile)
	
	#Import Propid results to Database
	if mode == 1:
		print("Saving PropIDs and GURAS Data to Database...")
		print("------------------------------------------------")
		
		connection = connectDB()
		c = connection.cursor()
		
		#Start Bulk insert query
		iQuery = ''
		
		#Import PropIDs
		pNextId = getNextId("GURAS_lot_id","GURAS_Lot")
		
		for i, row in df_propidFinal.iterrows():
			if pd.notna(row['attributes.sppropid']):
				unique_propid = int(row['attributes.sppropid'])
				i_sppropid = int(row['attributes.sppropid'])
			else:
				if pd.notna(row['attributes.propid']):
					unique_propid = int(row['attributes.propid'])
				else:
					unique_propid = 'null'
				i_sppropid = 'null'
				
			iQuery += " INTO GURAS_Lot VALUES ({},{},{},{},'{}','{}','{}','{}',{},'{}',CURRENT_TIMESTAMP)".format(pNextId,ifnull(row['attributes.propid'],'null'),i_sppropid,ifnull(unique_propid,'null'),row['ptlotsecpn'],row['lot_no'],row['section_no'],row['plan_type'],row['plan_no'],sys.argv[1])
			
			pNextId += 1
			
			#Every 1000 commit to database
			if (i + 1) % 200 == 0 or (i + 1) == len(df_propidFinal):
				#print("INSERT ALL {} SELECT 1 FROM DUAL".format(iQuery))
				c.execute("INSERT ALL {} SELECT 1 FROM DUAL".format(iQuery))
				iQuery = ''
		
		print("{} Lot-PropID records inserted into GURAS_Lot table!".format(i + 1))
		
		#Import GURAS Data
		gNextId = getNextId("GURAS_data_id","GURAS_Data")
		iQuery = ''
		
		for i, row in df_gurasResults.iterrows():
			if pd.notna(row['attributes.sppropid']):
				unique_propid = int(row['attributes.sppropid'])
				i_sppropid = int(row['attributes.sppropid'])
			else:
				if pd.notna(row['attributes.propid']):
					unique_propid = int(row['attributes.propid'])
				else:
					unique_propid = 'null'
				i_sppropid = 'null'
			
			#Convert Epoch/Unix Time
			i_createdate = datetime.fromtimestamp(row['attributes.createdate']/1000)
			
			iQuery += " INTO GURAS_Data VALUES ({},'{}','{}','{}',{},{},{},{},{},{},{},{},{},{},'{}',{},'{}','{}',{},'{}','{}','{}','{}','{}','{}',{},'{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}',{},{},'{}',{},'{}',{},{},'{}',{},{},{},{},{},{},{},'{}',CURRENT_TIMESTAMP)".format(gNextId,row['merged_property_description'].replace("'","''"),row['merged_address'].replace("'","''"),row['merged_suburb'].replace("'","''"),row['attributes.objectid'],"TO_DATE('{}', 'yyyy-mm-dd hh24:mi:ss')".format(i_createdate),row['attributes.gurasid'],row['attributes.addresstype'],row['attributes.ruraladdress'],row['attributes.principaladdresstype'],row['attributes.addressstringtype'],row['attributes.principaladdresssiteoid'],row['attributes.officialaddressstringoid'],row['attributes.roadside'],ifnull(row['attributes.housenumberfirstprefix'],''),ifnull(row['attributes.housenumberfirst'],'null'),ifnull(row['attributes.housenumberfirstsuffix'],''),ifnull(row['attributes.housenumbersecondprefix'],''),ifnull(row['attributes.housenumbersecond'],'null'),ifnull(row['attributes.housenumbersecondsuffix'],''),ifnull(row['attributes.roadname'],'').replace("'","''"),ifnull(row['attributes.roadtype'],''),ifnull(row['attributes.roadsuffix'],''),ifnull(row['attributes.unittype'],''),ifnull(row['attributes.unitnumberprefix'],''),ifnull(row['attributes.unitnumber'],'null'),ifnull(row['attributes.unitnumbersuffix'],''),ifnull(row['attributes.leveltype'],''),ifnull(row['attributes.levelnumberprefix'],''),ifnull(row['attributes.levelnumber'],''),ifnull(row['attributes.levelnumbersuffix'],''),ifnull(row['attributes.addresssitename'],'').replace("'","''"),ifnull(row['attributes.buildingname'],'').replace("'","''"),ifnull(row['attributes.locationdescription'],'').replace("'","''"),ifnull(row['attributes.privatestreetname'],'').replace("'","''"),ifnull(row['attributes.privatestreettype'],''),ifnull(row['attributes.privatestreetsuffix'],''),ifnull(row['attributes.secondroadname'],'').replace("'","''"),ifnull(row['attributes.secondroadtype'],''),ifnull(row['attributes.secondroadsuffix'],''),row['attributes.suburbname'].replace("'","''"),row['attributes.state'],row['attributes.postcode'],row['attributes.council'],ifnull(row['attributes.deliverypointid'],'null'),ifnull(row['attributes.deliverypointbarcode'],''),ifnull(row['attributes.addressconfidence'],'null'),ifnull(row['attributes.contributororigin'],'null'),ifnull(row['attributes.contributorid'],''),ifnull(row['attributes.contributoralignment'],'null'),ifnull(row['attributes.routeoid'],'null'),ifnull(row['attributes.gnafprimarysiteid'],'null'),ifnull(row['attributes.containment'],'null'),row['attributes.propid'],i_sppropid,unique_propid,sys.argv[1])
			
			gNextId += 1
			
			#Every 1000 commit to database
			if (i + 1) % 200 == 0 or (i + 1) == len(df_gurasResults):
				c.execute("INSERT ALL {} SELECT 1 FROM DUAL".format(iQuery))
				iQuery = ''
				
		print("{} GURAS records inserted into GURAS_Data table!".format(i + 1))
		print("------------------------------------------------")
		
		c.execute("commit")
		connection.close()
		
	print("GURAS Extract Process completed!!")
	print("------------------------------------------------")
		


